# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.

import openpyxl
import smtplib
import sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('.\\files\\duesRecords.xlsx')
sheet = wb.worksheets[0]

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.google.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('email_here', sys.argv[1])

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\n\nDear %s,\nRecords indicate that you have not paid dues for %s. Please pay " \
           "ASAP! Thanks!" % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('myemail@address.com', email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()